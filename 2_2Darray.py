from openpyxl import load_workbook
import sys

# 함수 생성
def save_value(value):
    print(value)
    # value , 기준으로 배열로 만들기
    
    print(value)
    # 엑셀 파일 불러오기
    wb = load_workbook('test.xlsx')
    wsArr = wb.sheetnames
    ws = wb[wsArr[0]]
    print((len(value)))
    for i in range(1,len(value)):
        print(i)
        d = value[i].split(',')
        ws.append(d)
   
    
    wb.save('test3.xlsx')
    # 엑셀 파일 닫기
    wb.close()
    # 셀의 값을 반환
    return 


if __name__ == '__main__':
    save_value(sys.argv)