from openpyxl import load_workbook
import sys
import json
# 함수 생성
def save_value(value):
    
    # value , 기준으로 배열로 만들기
    print('=======================1')
    # string 을 json 으로 변환
    value = json.loads(value)
    print(value)
    print('=======================2')
    # 엑셀 파일 불러오기
    wb = load_workbook('test.xlsx')
    wsArr = wb.sheetnames
    ws = wb[wsArr[0]]
    print('=======================3')
    
    for arr in value:
        print(arr)
        for key,val in arr.items():
            print(key,val)
            
        
        
        
   
    
    wb.save('test4.xlsx')
    # 엑셀 파일 닫기
    wb.close()
    # 셀의 값을 반환
    return 


if __name__ == '__main__':
    save_value(sys.argv[1])