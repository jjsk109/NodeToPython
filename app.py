from openpyxl import load_workbook
import sys

# 함수 생성
def save_value(x, y, value):
    # 엑셀 파일 불러오기
    wb = load_workbook('test.xlsx')
    wsArr = wb.sheetnames
    print(wsArr)
    ws = wb[wsArr[0]]
    ws.cell(row=x, column=y, value=value)
    wb.save('test2.xlsx')
    # 엑셀 파일 닫기
    wb.close()
    # 셀의 값을 반환
    return 

print( __name__)    # __main__ 출력
if __name__ == '__main__':
    save_value(sys.argv[1], sys.argv[2],sys.argv[2])